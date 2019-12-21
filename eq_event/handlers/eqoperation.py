# coding: utf-8
import time
import application
import json
import tornado.web
import threading
import os
import sys
reload(sys)
sys.setdefaultencoding("utf-8")
from openpyxl import load_workbook, Workbook
from multiprocessing import Process
from dal.dal_earthquake import Dal_Earthquake
from dal.dal_fj import Dal_Fj
from handlers.editInfo import get_FJ_by_eqID
from model.earthquake import Earthquake
from configs.logger import LOG


class EqDetail(tornado.web.RequestHandler):
    """
    此接口和编辑地震信息接口一样，位置不该放这儿，为了方便暂且放这儿后期可优化
    """
    def get(self):
        # query earthquake info
        earthquake_id = self.get_argument('earthquake_id', 0)
        response = {'status': 200, 'message': 'ok'}
        if not earthquake_id:
            response = {'status': 1005, 'message': 'earthquake_id not null'}
            # not exist is add info
            # return None
        else:
            earthquake_info = Dal_Earthquake().getEarthquake(int(earthquake_id))
            response.update({"data": dict(earthquake_info)})
            response['data']['date'] = response['data']['date'].strftime('%Y-%m-%d %H:%M:%S') \
                if response['data']['date'] else response['data']['date']
            # 添加附件数据
            response['data']["attachment"] = get_FJ_by_eqID(earthquake_id)

        # response = json.dumps(response)
        # response = self.get_argument('jsoncallback')+"("+response+")"
        self.set_header("Access-Control-Allow-Origin", "*")
        self.write(response)


class EqUpload(tornado.web.RequestHandler):
    """数据上传"""
    def post(self):
        post_data = {}
        response = {'status': 200, 'message': 'ok', 'data': {}}
        for key in self.request.arguments:
            post_data[key] = self.get_arguments(key)[0]

        root_path = application.SETTINGS['static_path']
        if not os.path.exists(root_path):
            os.mkdir(root_path)
        upload_path = os.path.join(root_path, "excel_upload")
        if not os.path.exists(upload_path):
            os.mkdir(upload_path)

        for upfile in self.request.files['file']:
                # currpath +'\static' 根路劲
                tmpfile = open(upload_path + "\\" + upfile['filename'], "wb")
                # tmpfile = open(upload_path, "wb")
                tmpfile.write(upfile['body'])
                tmpfile.flush()
                tmpfile.close()
                # self.addUserFile(post_data['id'], str(upfile['filename']))
                # 附件插入数据库
                time.sleep(1)
                # 开始读取数据
        # return
        # response = json.dumps(response)
        # response = self.get_argument('jsoncallback')+"("+response+")"
        self.set_header("Access-Control-Allow-Origin", "*")
        self.write(response)
        return None


class EqDown(tornado.web.RequestHandler):
    """地震数据删除（接口已经删除）"""
    def post(self):
        post_data = {}
        response = {'status': 200, 'message': 'ok', 'data': {}}
        for key in self.request.arguments:
            post_data[key] = self.get_arguments(key)[0]

        # userRequest = 12
        if not post_data["earthquake_id"] and not post_data["attachment_name"]:
            response = {'status': 10004, 'message': "earthquake_id and attachment_name not null"}
        else:
            # 查询附件表中是否存在
            result = Dal_Fj.query("fj", eqid=post_data["earthquake_id"], fjname=post_data["attachment_name"])
            if not result:
                response = {'status': 10004, 'message': "delete fj not exist.."}
            else:
                # 假设移除成功..
                Dal_Fj().remove("fj",  eqid=post_data["earthquake_id"], fjname=post_data["attachment_name"])
        # response = json.dumps(response)
        # response = self.get_argument('jsoncallback')+"("+response+")"
        self.set_header("Access-Control-Allow-Origin", "*")
        self.write(response)
        return None


def is_exist_file(path):
    count = 0
    while True:
        if count > 5:
            return False
        else:
            if not os.path.exists(path):
                count += 1
                continue

            return True


def read_excel_data():
    """
    单独启动一个进程监听上传的excel表格，一旦有上传的数据就开始读，读完之后删除
    @return:
    """
    # 目录已经确定
    # 程序一旦启动，就生成下载数据的模板
    root_path = os.getcwd()
    # print root_path
    upload_path = os.path.join(root_path, "static")
    if not os.path.exists(upload_path):
        os.mkdir(upload_path)
    create_excel(upload_path)

    excel_path = os.path.join(upload_path, "excel_upload")
    while True:
        if os.path.exists(excel_path):
            excel_list = os.listdir(excel_path)
            if not excel_list:
                continue
            LOG.info("[+] file upload is reading..")
            for excel in excel_list:

                wb = load_workbook(os.path.join(excel_path, excel))
                sheet = wb.active
                for row in list(sheet.rows)[1:]:
                    print row[4].value
                    # timestamp = Utils().string2timestamp(row[4].value)
                    now = row[4].value
                    newEarth = Earthquake(eqname=row[0].value,
                                          longitude=row[1].value,
                                          latitude=row[2].value,
                                          level=row[3].value,
                                          depth=row[5].value,
                                          introduction=row[6].value,
                                          date=now.date(),
                                          Year=now.year,
                                          Month=now.month,
                                          Day=now.day,
                                          hour=now.hour,
                                          minute=now.minute,
                                          second=now.second)
                    newID = Dal_Earthquake().addEarthquake(newEarth)

                os.remove(os.path.join(excel_path, excel))
            LOG.info("[+] file upload success and start create excel..")
            # 一旦发现有上传，就立即更新下载的模板
            create_excel(upload_path)


def create_excel(upload_path):
    """
    生成excel表格用于下载，只有在用户上传数据的时候下载的表格曹辉更新
    @param upload_path:
    @return:
    """
    # 程序一旦启动，就生成下载数据的模板
    # root_path = os.getcwd()
    # upload_path = root_path.replace("handlers", "static")
    down_excel_path = os.path.join(upload_path, "excel_down")
    if not os.path.exists(down_excel_path):
        os.mkdir(down_excel_path)

    dowm_excel_file = os.path.join(down_excel_path, "earthquake.xlsx")
    # 如果文件纯在 先删除
    if os.path.exists(dowm_excel_file):
        os.remove(dowm_excel_file)
        LOG.info("[+] delete old excel..")

    wb = Workbook()
    sheet = wb.active

    kward={}
    eq_list = Dal_Earthquake().query('earthquake', **kward)
    # 添加标题
    eq_title = ["eqname",  "longitude", "latitude", "level", "date", "depth", "attachment", "introduction"]
    sheet.append(eq_title)
    # 为每个地震数据添加附件  每个地震数据是字典
    for eq in eq_list:
        eq_field = list()
        # 字段排序 名称（eqname） 经度(longitude) 纬度(latitude) 级别（level） 日期(date)
        # 地震深度(depth)  附加内容()  说明(introduction)
        eq_field = eq_field + [eq.get("eqname"), eq.get("longitude"), eq.get("latitude"),
                               eq.get("level"), eq.get("date"), eq.get("depth")]
        #
        fj_centent = json.loads(get_FJ_by_eqID(eq.get("id")))
        fjname_list = [fj.get("fjname", "") for fj in fj_centent]
        eq_field.append(",".join(fjname_list))
        eq_field.append(eq.get("introduction"))
        sheet.append(eq_field)
    wb.save(dowm_excel_file)
    LOG.info("[+] create excel is success..")

# read_excel_data()

re = threading.Thread(target=read_excel_data)
re.start()
if __name__ == '__main__':
    # aa()
    pass









